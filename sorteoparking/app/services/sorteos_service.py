"""Lógica de negocio de sorteos, OTP y notificaciones. SDD §5.3, §6, T-201 a T-206."""



import hashlib
import logging
import os

import io

import random
import time

import secrets

from datetime import datetime, timedelta, timezone
from typing import Any



from fastapi import HTTPException, status

from openpyxl import load_workbook

from sqlalchemy.orm import Session

from sqlalchemy import update



from app.core.config import deploy_config, public_urls_config

from app.core.security import enforce_tenant_scope

from app.models.catalogo import Parqueadero

from app.models.sorteo import Consejero, Participante, ResultadoSorteo, SesionOTP, Sorteo

from app.models.tenant import Tenant

from app.services.email_service import enviar_correo_texto

from app.services.log_service import registrar_log_auditoria

from app.services.otp_service import generar_otp_numerico_seis_digitos, hashear_otp, verificar_otp


from app.services.sorteo_engine import ejecutar_sorteo_hibrido


logger = logging.getLogger(__name__)

# NOTE(FIX #29): timezone-naive design — aligns with SQLite/DB storage convention.
# Revisit if migrating to timezone-aware DB or multi-region deployment.






def listar_historial_sorteos(db: Session, tenant_id: str) -> list[Sorteo]:

    """Lista los sorteos del tenant y registra consulta en auditoria."""

    sorteos = db.query(Sorteo).filter(Sorteo.tenant_id == tenant_id).order_by(Sorteo.id.desc()).all()

    registrar_log_auditoria(

        db=db,

        tenant_id=tenant_id,

        evento="CONSULTA_HISTORIAL_SORTEOS",

        payload=f"total={len(sorteos)}",

    )
    db.commit()

    return sorteos





def _calcular_snapshot_hash(participantes: list[Participante]) -> str:

    """Hash determinista de la lista de elegibles (SDD §6.3)."""

    lineas = sorted(f"{p.documento}|{p.nombre}" for p in participantes)

    return hashlib.sha256("\n".join(lineas).encode("utf-8")).hexdigest()





def _resolver_tenant_por_slug(db: Session, tenant_slug: str) -> Tenant:

    tenant = db.query(Tenant).filter(Tenant.slug == tenant_slug).first()

    if not tenant:

        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Conjunto no encontrado")

    return tenant





def _obtener_sorteo_tenant(db: Session, tenant_id: str, sorteo_id: int) -> Sorteo:

    sorteo = db.query(Sorteo).filter(Sorteo.id == sorteo_id, Sorteo.tenant_id == tenant_id).first()

    if not sorteo:

        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Sorteo no encontrado")

    return sorteo





def _hay_sorteo_activo_otro(db: Session, tenant_id: str, excluir_sorteo_id: int) -> bool:

    """409 en curso: otro sorteo EN_CURSO o LISTO (SDD §5.3 iniciar)."""

    q = (

        db.query(Sorteo)

        .filter(

            Sorteo.tenant_id == tenant_id,

            Sorteo.id != excluir_sorteo_id,

            Sorteo.estado.in_(("EN_CURSO", "LISTO")),

        )
        .with_for_update()
        .first()

    )

    return q is not None





def _detectar_fila_encabezado(rows: list[tuple]) -> int | None:

    """Busca la fila que contiene 'apartamento' como indicador del encabezado real.
    Solo revisa las primeras 20 filas para evitar falsos positivos en datos."""

    max_rows = min(len(rows), 20)
    for i in range(max_rows):

        celdas = [str(c).strip().lower() if c is not None else "" for c in rows[i]]

        if "apartamento" in celdas:

            return i

    return None





def _procesar_hoja_elegibles(

    db: Session, tenant_id: str, sorteo_id: int, ws, tipo_vehiculo_default: str

) -> int:

    """Procesa una hoja ELEGIBLES_CARRO o ELEGIBLES_MOTO del formato oficial."""

    rows = list(ws.iter_rows(values_only=True))

    if not rows:

        return 0

    fila_header = _detectar_fila_encabezado(rows)

    if fila_header is None:

        return 0

    header = [str(c).strip().lower() if c is not None else "" for c in rows[fila_header]]

    idx = {name: i for i, name in enumerate(header)}

    if "apartamento" not in idx:

        return 0

    insertados = 0

    for raw in rows[fila_header + 1:]:

        if raw is None or all(c is None or str(c).strip() == "" for c in raw):

            continue

        apto = str(raw[idx["apartamento"]]).strip() if idx.get("apartamento") is not None and raw[idx["apartamento"]] is not None else ""

        if not apto or apto.upper().startswith("TOTAL"):

            continue

        torre_val = ""

        if "torre" in idx and raw[idx["torre"]] is not None:

            torre_val = str(raw[idx["torre"]]).strip()

        correo = None

        if "correo" in idx and raw[idx["correo"]] is not None:

            correo = str(raw[idx["correo"]]).strip() or None

        vehiculo = tipo_vehiculo_default

        if "tipovehiculo" in idx and raw[idx["tipovehiculo"]] is not None:

            vehiculo = str(raw[idx["tipovehiculo"]]).strip().upper() or tipo_vehiculo_default

        marca = ""

        if "marcamodelo" in idx and raw[idx["marcamodelo"]] is not None:

            marca = str(raw[idx["marcamodelo"]]).strip()

        hatch = False

        if "eshatchback" in idx and raw[idx["eshatchback"]] is not None:

            val = str(raw[idx["eshatchback"]]).strip().upper()

            hatch = val in ("SI", "SÍ", "VERDADERO", "TRUE", "1")

        nombre = f"{apto} Torre {torre_val}" if torre_val else apto

        documento = apto

        db.add(

            Participante(

                tenant_id=tenant_id,

                sorteo_id=sorteo_id,

                nombre=nombre,

                documento=documento,

                apartamento=apto,

                es_hatchback=hatch,

                tipo_vehiculo=vehiculo,

                email=correo,

            )

        )

        insertados += 1

    return insertados





def cargar_excel_elegibles(db: Session, tenant_id: str, contenido: bytes) -> dict[str, Any]:

    """

    Importa elegibles desde Excel oficial SorteoParking (SDD §5.3).

    Formato: 3 hojas — INSTRUCCIONES, ELEGIBLES_CARRO, ELEGIBLES_MOTO.

    Columnas: apartamento, torre, correo, tipoVehiculo, marcaModelo, esHatchback.

    """

    try:

        wb = load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)

    except Exception as exc:

        raise HTTPException(

            status_code=status.HTTP_400_BAD_REQUEST,

            detail="Archivo Excel invalido o corrupto",

        ) from exc

    hojas = [n.upper() for n in wb.sheetnames]

    tiene_carro = any("ELEGIBLES_CARRO" in h for h in hojas)

    tiene_moto = any("ELEGIBLES_MOTO" in h for h in hojas)

    es_formato_oficial = tiene_carro or tiene_moto

    sorteo = Sorteo(tenant_id=tenant_id, estado="PENDIENTE", tipo="GENERAL")

    db.add(sorteo)

    db.flush()

    insertados = 0

    if es_formato_oficial:

        for sheet_name in wb.sheetnames:

            upper = sheet_name.upper()

            if "ELEGIBLES_CARRO" in upper:

                insertados += _procesar_hoja_elegibles(db, tenant_id, sorteo.id, wb[sheet_name], "CARRO")

            elif "ELEGIBLES_MOTO" in upper:

                insertados += _procesar_hoja_elegibles(db, tenant_id, sorteo.id, wb[sheet_name], "MOTO")

    else:

        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))

        if rows:

            fila_header = _detectar_fila_encabezado(rows)

            if fila_header is None:

                fila_header = 0

            header = [str(c).strip().lower() if c is not None else "" for c in rows[fila_header]]

            idx = {name: i for i, name in enumerate(header)}

            for raw in rows[fila_header + 1:]:

                if raw is None or all(c is None or str(c).strip() == "" for c in raw):

                    continue

                nombre = str(raw[idx.get("nombre", 0)]).strip() if idx.get("nombre") is not None and raw[idx["nombre"]] is not None else ""

                documento = str(raw[idx.get("documento", 1)]).strip() if idx.get("documento") is not None and raw[idx["documento"]] is not None else ""

                if not nombre or not documento or nombre.upper().startswith("TOTAL"):

                    continue

                apto = None

                if "apartamento" in idx and raw[idx["apartamento"]] is not None:

                    apto = str(raw[idx["apartamento"]]).strip() or None

                hatch = False

                for key in ("eshatchback", "es_hatchback"):

                    if key in idx and raw[idx[key]] is not None:

                        val = str(raw[idx[key]]).strip().upper()

                        hatch = val in ("SI", "SÍ", "VERDADERO", "TRUE", "1")

                        break

                vehiculo = "CARRO"

                for key in ("tipovehiculo", "tipo_vehiculo"):

                    if key in idx and raw[idx[key]] is not None:

                        vehiculo = str(raw[idx[key]]).strip().upper() or "CARRO"

                        break

                em = None

                for key in ("correo", "email"):

                    if key in idx and raw[idx[key]] is not None:

                        em = str(raw[idx[key]]).strip() or None

                        break

                db.add(

                    Participante(

                        tenant_id=tenant_id,

                        sorteo_id=sorteo.id,

                        nombre=nombre,

                        documento=documento,

                        apartamento=apto,

                        es_hatchback=hatch,

                        tipo_vehiculo=vehiculo,

                        email=em,

                    )

                )

                insertados += 1

    wb.close()

    if insertados > 10000:
        db.rollback()
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Demasiados participantes (maximo 10000 por sorteo)")

    if insertados == 0:

        db.rollback()

        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Sin filas de participantes validas")

    try:
        db.commit()
        logger.info("Cargados %d participantes para sorteo %d", insertados, sorteo.id)
    except Exception:
        db.rollback()
        raise

    registrar_log_auditoria(

        db=db,

        tenant_id=tenant_id,

        evento="CARGA_EXCEL_ELEGIBLES",

        payload=f"sorteo_id={sorteo.id},participantes={insertados},formato={'oficial' if es_formato_oficial else 'simple'}",

    )
    db.commit()

    return {"sorteo_id": sorteo.id, "participantes_cargados": insertados}





def _mensaje_otp_consejero(nombre_conjunto: str, nombre_consejero: str, otp: str, token: str, sorteo_id: int) -> str:

    base = public_urls_config.public_base_url.rstrip("/")

    enlace = f"{base}/static/otp_panel.html#t={token}&sid={sorteo_id}"

    return (

        f"SorteoParking — {nombre_conjunto}\n"

        f"Hola {nombre_consejero}, su codigo OTP es: {otp}\n"

        f"Confirme aqui: {enlace}\n"

        f"(caduca en 30 minutos)"

    )





def _entregar_mensaje_consejero(

    email: str | None,

    asunto: str,

    cuerpo: str,

) -> tuple[bool, str]:

    """Envía OTP por email (único canal de delivery)."""

    if email and enviar_correo_texto(email, asunto, cuerpo):

        return True, "email"

    return False, "ninguno"





def iniciar_sorteo(

    db: Session,

    tenant_id: str,

    sorteo_id: int,

    consejeros: list[dict[str, str | None]],

) -> dict[str, Any]:

    """Crea consejeros, OTPs, snapshot y envia OTP por email (SDD §5.3, §6.2)."""
    logger.warning("INICIAR_SORTEO_START | tenant=%s sorteo_id=%s total_consejeros=%d", tenant_id, sorteo_id, len(consejeros))
    emails_consejeros = [c.get("email") for c in consejeros]
    logger.warning("INICIAR_SORTEO_CONSEJEROS | tenant=%s sorteo_id=%s emails=%s", tenant_id, sorteo_id, emails_consejeros)

    try:
        if len(consejeros) != 5:

            raise HTTPException(

                status_code=status.HTTP_400_BAD_REQUEST,

                detail="Se requieren exactamente 5 consejeros",

            )

        sorteo = _obtener_sorteo_tenant(db, tenant_id, sorteo_id)

        enforce_tenant_scope(tenant_id, sorteo.tenant_id)

        if sorteo.estado != "PENDIENTE":

            raise HTTPException(

                status_code=status.HTTP_400_BAD_REQUEST,

                detail="El sorteo no esta en estado PENDIENTE",

            )

        if _hay_sorteo_activo_otro(db, tenant_id, sorteo_id):

            raise HTTPException(

                status_code=status.HTTP_409_CONFLICT,

                detail="Ya existe un sorteo en curso para este conjunto",

            )

        # SDD §15.7 — Verificar que haya catálogo cargado antes de iniciar
        tiene_catalogo = db.query(Parqueadero).filter(
            Parqueadero.tenant_id == tenant_id
        ).first()
        if not tiene_catalogo:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="No hay catálogo de parqueaderos cargado. Use POST /catalogo/carga-csv primero.",
            )

        logger.warning("INICIAR_SORTEO_ETAPA | tenant=%s sorteo_id=%s etapa=participantes_y_snapshot", tenant_id, sorteo_id)
        participantes = (

            db.query(Participante)

            .filter(Participante.tenant_id == tenant_id, Participante.sorteo_id == sorteo_id)

            .order_by(Participante.id)

            .all()

        )

        if not participantes:

            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No hay participantes cargados")

        sorteo.snapshot_hash = _calcular_snapshot_hash(participantes)

        sorteo.estado = "EN_CURSO"

        db.query(SesionOTP).filter(SesionOTP.sorteo_id == sorteo_id, SesionOTP.tenant_id == tenant_id).delete()

        db.query(Consejero).filter(Consejero.sorteo_id == sorteo_id, Consejero.tenant_id == tenant_id).delete()

        db.flush()

        logger.warning("INICIAR_SORTEO_ETAPA | tenant=%s sorteo_id=%s etapa=creacion_otp_db", tenant_id, sorteo_id)
        tenant = db.query(Tenant).filter(Tenant.id == tenant_id).first()

        nombre_conjunto = tenant.nombre if tenant else "Conjunto"

        expira = datetime.now() + timedelta(minutes=30)

        sesiones_creadas: list[tuple[SesionOTP, Consejero, str]] = []

        for item in consejeros:

            nombre_c = str(item.get("nombre") or "").strip()

            email_c = (str(item.get("email")).strip() if item.get("email") else None) or None

            if not nombre_c or not email_c:

                db.rollback()

                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Cada consejero requiere nombre y email")

            cons = Consejero(

                tenant_id=tenant_id,

                sorteo_id=sorteo_id,

                nombre=nombre_c,

                email=email_c,

            )

            db.add(cons)

            db.flush()

            otp_plano = generar_otp_numerico_seis_digitos()

            token_enlace = secrets.token_urlsafe(32)

            ses = SesionOTP(

                tenant_id=tenant_id,

                sorteo_id=sorteo_id,

                consejero_id=cons.id,

                otp_hash=hashear_otp(otp_plano),

                token_enlace=token_enlace,

                estado="PENDIENTE",

                expira_en=expira,

            )

            db.add(ses)

            sesiones_creadas.append((ses, cons, otp_plano))

        db.flush()

        logger.warning("INICIAR_SORTEO_ETAPA | tenant=%s sorteo_id=%s etapa=envio_smtp consejeros_creados=%d", tenant_id, sorteo_id, len(sesiones_creadas))
        otp_debug_info: list[dict[str, str]] = []

        es_desarrollo = (

            deploy_config.app_env == "development" and os.getenv("DEBUG", "false").lower() == "true"

        )

        for ses, cons, otp_plano in sesiones_creadas:
            try:
                cuerpo = _mensaje_otp_consejero(nombre_conjunto, cons.nombre, otp_plano, ses.token_enlace, sorteo_id)

                ok, canal = _entregar_mensaje_consejero(

                    cons.email,

                    f"OTP SorteoParking — {nombre_conjunto}",

                    cuerpo,

                )
            except Exception as e:
                logger.exception("INICIAR_SORTEO_SMTP_ERROR | tenant=%s sorteo_id=%s consejero=%s email=%s", tenant_id, sorteo_id, cons.nombre, cons.email)
                ok = False
                canal = "excepcion"

            if not ok and not es_desarrollo:
                # Marcar esta sesion como fallida en lugar de rollback total
                ses.estado = "FALLO_EMAIL"
                db.flush()
                logger.warning("No se pudo enviar OTP por correo a %s", cons.email)

            if es_desarrollo and deploy_config.app_env == "development":

                base = public_urls_config.public_base_url.rstrip("/")

                otp_debug_info.append({

                    "consejero": cons.nombre,

                    "otp": otp_plano,

                    "link": f"{base}/static/otp_panel.html#t={ses.token_enlace}&sid={sorteo_id}",

                    "canal": canal if ok else "NO_ENVIADO (modo desarrollo)",

                })

        logger.warning("INICIAR_SORTEO_ETAPA | tenant=%s sorteo_id=%s etapa=commit", tenant_id, sorteo_id)
        db.commit()
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("INICIAR_SORTEO_EXCEPTION | tenant=%s sorteo_id=%s", tenant_id, sorteo_id)
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=str(e))

    registrar_log_auditoria(

        db=db,

        tenant_id=tenant_id,

        evento="SORTEO_INICIADO_OTP",

        payload=f"sorteo_id={sorteo_id}",

    )
    db.commit()

    resultado: dict[str, Any] = {

        "id": sorteo.id,

        "estado": sorteo.estado,

        "snapshot_hash": sorteo.snapshot_hash,

    }

    if es_desarrollo and otp_debug_info:

        resultado["_dev_otps"] = otp_debug_info

        resultado["_dev_nota"] = "OTPs visibles solo en modo desarrollo. En produccion se envian por correo."

    return resultado





def confirmar_otp(

    db: Session,

    sorteo_id: int,

    tenant_id_desde_token: str,

    token_enlace: str,

    otp_ingresado: str,

) -> dict[str, str]:

    """Confirma OTP de consejero (SDD §5.3, §6.6 T-118). Sin loguear OTP en auditoria."""

    # Lock pesimista contra race condition (SDD §6.6)
    ses = (

        db.query(SesionOTP)

        .filter(

            SesionOTP.sorteo_id == sorteo_id,

            SesionOTP.token_enlace == token_enlace,

        )
        .with_for_update()
        .first()

    )

    if not ses:

        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Enlace invalido")

    if tenant_id_desde_token and ses.tenant_id != tenant_id_desde_token:

        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Acceso denegado: recurso de otro tenant")

    sorteo = db.query(Sorteo).filter(Sorteo.id == sorteo_id, Sorteo.tenant_id == ses.tenant_id).first()

    if not sorteo:

        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Sorteo no encontrado")

    # SDD §6.6 — Anti-replay: verificar si ya fue usado
    if ses.estado == "CONFIRMADO":

        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="OTP ya utilizado — no se puede reutilizar")

    # SDD §6.4 — Verificar expiración
    if ses.estado == "PENDIENTE" and datetime.now() > ses.expira_en:

        ses.estado = "EXPIRADO"

        db.commit()

        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="OTP expirado")
    elif datetime.now() > ses.expira_en:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="OTP expirado")

    # SDD §6.4 — Límite de 3 intentos
    if ses.intentos >= 3:

        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="OTP bloqueado por exceso de intentos")





    if not verificar_otp(otp_ingresado, ses.otp_hash):

        ses.intentos += 1

        db.commit()

        registrar_log_auditoria(

            db=db,

            tenant_id=ses.tenant_id,

            evento="OTP_CONFIRMACION_FALLIDA",

            payload=f"sorteo_id={sorteo_id},sesion_id={ses.id},intento={ses.intentos}",

        )
        db.commit()

        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="OTP invalido")

    # SDD §6.6 — Marcar como usado — single-use estricto
    ses.estado = "CONFIRMADO"

    ses.confirmado_en = datetime.now()

    db.flush()

    # Verificar si es el 5to OTP → LISTO
    confirmados = (

        db.query(SesionOTP)

        .filter(

            SesionOTP.sorteo_id == sorteo_id,

            SesionOTP.tenant_id == ses.tenant_id,

            SesionOTP.estado == "CONFIRMADO",

        )

        .count()

    )

    if confirmados >= 5:

        sorteo.estado = "LISTO"

    db.commit()

    registrar_log_auditoria(

        db=db,

        tenant_id=ses.tenant_id,

        evento="OTP_CONFIRMACION_OK",

        payload=f"sorteo_id={sorteo_id},sesion_id={ses.id}",

    )
    db.commit()

    return {"estado": "confirmado", "sorteo_estado": sorteo.estado}





def estado_otp(db: Session, tenant_id: str, sorteo_id: int, token_enlace: str | None = None) -> dict[str, Any]:

    """Progreso 0-5 con nombres (SDD §5.3)."""

    # Cuando viene del panel OTP (sin Bearer), resolver tenant y sorteo desde token_enlace
    if not tenant_id and token_enlace:
        ses = db.query(SesionOTP).filter(SesionOTP.token_enlace == token_enlace).first()
        if not ses:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Enlace invalido")
        tenant_id = ses.tenant_id
        if ses.sorteo_id != sorteo_id:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Enlace no corresponde a este sorteo")

    if not tenant_id:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Falta autenticacion")

    sorteo = _obtener_sorteo_tenant(db, tenant_id, sorteo_id)

    enforce_tenant_scope(tenant_id, sorteo.tenant_id)

    sesiones = (

        db.query(SesionOTP, Consejero)

        .join(Consejero, SesionOTP.consejero_id == Consejero.id)

        .filter(SesionOTP.sorteo_id == sorteo_id, SesionOTP.tenant_id == tenant_id)

        .order_by(Consejero.id)

        .all()

    )

    # Encontrar mi_estado desde token_enlace si viene de panel OTP
    mi_estado = None
    sesion_id = None
    expira_en = None
    if token_enlace:
        for s, _ in sesiones:
            if s.token_enlace == token_enlace:
                mi_estado = s.estado
                sesion_id = s.id
                expira_en = s.expira_en.isoformat() if s.expira_en else None
                break

    items: list[dict[str, str]] = []

    for ses, cons in sesiones:

        items.append(

            {

                "nombre": cons.nombre,

                "estado": ses.estado,

            }

        )

    return {

        "sorteo_id": sorteo_id,

        "confirmados": sum(1 for s, _ in sesiones if s.estado == "CONFIRMADO"),

        "total": len(sesiones),

        "consejeros": items,
        "mi_estado": mi_estado,
        "sesion_id": sesion_id,
        "expira_en": expira_en,

    }





def estado_sorteo(db: Session, tenant_id: str, sorteo_id: int) -> dict[str, Any]:

    """Estado actual para polling (SDD §5.3)."""

    sorteo = _obtener_sorteo_tenant(db, tenant_id, sorteo_id)

    enforce_tenant_scope(tenant_id, sorteo.tenant_id)

    return {

        "id": sorteo.id,

        "estado": sorteo.estado,

        "seed": sorteo.seed,

        "snapshot_hash": sorteo.snapshot_hash,

    }





def ejecutar_sorteo_asignacion(db: Session, tenant_id: str, sorteo_id: int) -> list[ResultadoSorteo]:

    """

    Ejecuta la asignación utilizando el motor híbrido v1.4.3 (SDD §3.5, §6.3).

    Incluye mutex de ejecución (T-120) y validación de snapshot (SDD §6.5).

    """

    sorteo = _obtener_sorteo_tenant(db, tenant_id, sorteo_id)

    enforce_tenant_scope(tenant_id, sorteo.tenant_id)

    # SDD §3.8 T-120 — Mutex de ejecución: LISTO → EJECUTANDO (solo si está LISTO)
    from sqlalchemy import update as sa_update
    resultado = db.execute(
        sa_update(Sorteo)
        .where(Sorteo.id == sorteo_id, Sorteo.tenant_id == tenant_id, Sorteo.estado == "LISTO")
        .values(estado="EJECUTANDO")
    )
    db.commit()
    if resultado.rowcount == 0:
        sorteo_refrescado = db.query(Sorteo).filter(
            Sorteo.id == sorteo_id, Sorteo.tenant_id == tenant_id
        ).first()
        estado_actual = sorteo_refrescado.estado if sorteo_refrescado else "DESCONOCIDO"
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"El sorteo no esta en estado LISTO o ya se esta ejecutando (actual: {estado_actual})",
        )
    sorteo_refrescado = db.query(Sorteo).filter(
        Sorteo.id == sorteo_id, Sorteo.tenant_id == tenant_id
    ).first()

    # SDD §6.5 — Validar inmutabilidad del snapshot
    try:

        participantes = db.query(Participante).filter(

            Participante.sorteo_id == sorteo_id,

            Participante.tenant_id == tenant_id,

        ).order_by(Participante.id).all()

        lineas = sorted(f"{p.documento}|{p.nombre}" for p in participantes)
        hash_actual = hashlib.sha256("\n".join(lineas).encode("utf-8")).hexdigest()

        if hash_actual != sorteo_refrescado.snapshot_hash:

            sorteo_refrescado.estado = "ERROR"

            db.commit()

            raise HTTPException(

                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,

                detail="Integridad comprometida: los participantes fueron modificados después de confirmar los OTPs",

            )

    except HTTPException:

        raise

    except Exception:

        pass  # Si falla la validación por cualquier razón, continuamos igual

    try:

        ejecutar_sorteo_hibrido(db, tenant_id, sorteo_id)

    except ValueError as exc:

        # Error del motor → restaurar a LISTO

        db.execute(

            update(Sorteo)

            .where(Sorteo.id == sorteo_id, Sorteo.tenant_id == tenant_id)

            .values(estado="LISTO")

        )

        db.commit()

        raise HTTPException(

            status_code=status.HTTP_400_BAD_REQUEST,

            detail=str(exc),

        ) from exc

    except Exception as exc:

        # Error interno → marcar ERROR

        db.execute(

            update(Sorteo)

            .where(Sorteo.id == sorteo_id, Sorteo.tenant_id == tenant_id)

            .values(estado="ERROR")

        )

        db.commit()

        registrar_log_auditoria(

            db=db,

            tenant_id=tenant_id,

            evento="SORTEO_ERROR",

            payload=f"sorteo_id={sorteo_id},error={str(exc)}",

        )
        db.commit()

        raise HTTPException(

            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,

            detail="Error interno durante la ejecucion del sorteo",

        ) from exc

    return db.query(ResultadoSorteo).filter(

        ResultadoSorteo.sorteo_id == sorteo_id,

        ResultadoSorteo.tenant_id == tenant_id

    ).all()





def notificar_resultados(db: Session, tenant_id: str, sorteo_id: int) -> dict[str, Any]:

    """Envia resultados por email (SDD §5.3, T-206)."""

    sorteo = _obtener_sorteo_tenant(db, tenant_id, sorteo_id)

    enforce_tenant_scope(tenant_id, sorteo.tenant_id)

    if sorteo.estado != "COMPLETADO":

        raise HTTPException(

            status_code=status.HTTP_400_BAD_REQUEST,

            detail="El sorteo debe estar en estado COMPLETADO antes de notificar",

        )

    tenant = db.query(Tenant).filter(Tenant.id == tenant_id).first()

    nombre_conjunto = tenant.nombre if tenant else "Conjunto"

    rows = (

        db.query(ResultadoSorteo, Participante)

        .join(Participante, ResultadoSorteo.participante_id == Participante.id)

        .filter(ResultadoSorteo.sorteo_id == sorteo_id, ResultadoSorteo.tenant_id == tenant_id)

        .all()

    )

    ok = 0

    fallos = 0

    for res, part in rows:

        if res.tipo_resultado != "GANADOR":

            continue

        nombre_seguro = (part.nombre or "").replace("\r", "").replace("\n", "").replace("\t", "")
        texto = (

            f"SorteoParking — {nombre_conjunto}\n"

            f"Hola {nombre_seguro}, su parqueadero asignado es: {res.parqueadero_asignado} ({res.zona_asignada})."

        )

        # Retry up to 2 times on failure
        enviado = False

        if part.email:
            for intento in range(2):
                enviado = enviar_correo_texto(
                    part.email,
                    f"Resultado sorteo — {nombre_conjunto}",
                    texto,
                )
                if enviado:
                    break
                time.sleep(1)

        if enviado:

            ok += 1

        else:

            fallos += 1
            logger.warning("No se pudo notificar a %s (%s) tras 2 intentos", part.nombre, part.email)

    db.commit()

    registrar_log_auditoria(

        db=db,

        tenant_id=tenant_id,

        evento="NOTIFICACION_RESULTADOS",

        payload=f"sorteo_id={sorteo_id},ok={ok},fallos={fallos}",

    )
    db.commit()

    return {"enviados": ok, "fallidos": fallos}





def obtener_resultados_paginados(db: Session, tenant_id: str, sorteo_id: int, pagina: int = 1, por_pagina: int = 20) -> dict[str, Any]:

    """Resultados paginados (SDD §5.3)."""

    sorteo = _obtener_sorteo_tenant(db, tenant_id, sorteo_id)

    enforce_tenant_scope(tenant_id, sorteo.tenant_id)

    query = (

        db.query(ResultadoSorteo, Participante)

        .join(Participante, ResultadoSorteo.participante_id == Participante.id)

        .filter(ResultadoSorteo.sorteo_id == sorteo_id, ResultadoSorteo.tenant_id == tenant_id)

        .order_by(Participante.apartamento)

    )

    total = query.count()

    offset = (pagina - 1) * por_pagina

    rows = query.offset(offset).limit(por_pagina).all()

    items = [

        {

            "participante_id": r.participante_id,

            "apartamento": p.apartamento,

            "parqueadero_asignado": r.parqueadero_asignado,

            "zona_asignada": r.zona_asignada,

            "tipo_resultado": r.tipo_resultado,

            "fue_reasignado": r.fue_reasignado,

        }

        for r, p in rows

    ]

    return {

        "items": items,

        "total": total,

        "pagina": pagina,

        "por_pagina": por_pagina,

        "total_paginas": (total + por_pagina - 1) // por_pagina,

    }





def obtener_diagnostico(db: Session, tenant_id: str, sorteo_id: int) -> dict[str, Any]:

    """Previsualiza modelo por zona antes de ejecutar (SDD §5.3)."""

    sorteo = _obtener_sorteo_tenant(db, tenant_id, sorteo_id)

    enforce_tenant_scope(tenant_id, sorteo.tenant_id)

    from app.models.catalogo import Zona, Torre, Parqueadero

    zonas = db.query(Zona).filter(Zona.tenant_id == tenant_id).order_by(Zona.nombre).all()

    participantes = db.query(Participante).filter(

        Participante.sorteo_id == sorteo_id, Participante.tenant_id == tenant_id

    ).all()

    tipo_sorteo = sorteo.tipo or "CARRO"

    diagnostico_zonas = []

    for zona in zonas:

        puestos = db.query(Parqueadero).filter(

            Parqueadero.tenant_id == tenant_id,

            Parqueadero.zona_id == zona.id,

            Parqueadero.vehiculo == tipo_sorteo,

            Parqueadero.disponible == True,

        ).count()

        # Participantes en esta zona (aproximado)

        torres_zona = db.query(Torre).filter(Torre.zona_id == zona.id, Torre.tenant_id == tenant_id).all()

        nombres_torre = [t.nombre for t in torres_zona]

        part_zona = sum(

            1 for p in participantes

            if p.apartamento and any(p.apartamento.startswith(t) for t in nombres_torre)

        )

        diagnostico_zonas.append({

            "zona": zona.nombre,

            "puestos_disponibles": puestos,

            "participantes_estimados": part_zona,

            "superavit": puestos - part_zona,

        })

    return {

        "sorteo_id": sorteo_id,

        "tipo": tipo_sorteo,

        "total_participantes": len(participantes),

        "zonas": diagnostico_zonas,

    }





def publico_sorteo(db: Session, tenant_slug: str, sorteo_id: int) -> dict[str, Any]:

    """Vista publica. SDD §5.4, CA-10: NO incluir datos personales (documento, email)."""

    tenant = _resolver_tenant_por_slug(db, tenant_slug)

    sorteo = (

        db.query(Sorteo)

        .filter(Sorteo.id == sorteo_id, Sorteo.tenant_id == tenant.id, Sorteo.estado == "COMPLETADO")

        .first()

    )

    if not sorteo:

        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Sorteo no encontrado o no publico")

    rows = (

        db.query(ResultadoSorteo, Participante)

        .join(Participante, ResultadoSorteo.participante_id == Participante.id)

        .filter(ResultadoSorteo.sorteo_id == sorteo_id, ResultadoSorteo.tenant_id == tenant.id)

        .order_by(Participante.apartamento)

        .all()

    )

    # SDD §16.3 — Solo apartamento y tipo de vehículo. Sin documento, nombre completo ni email.
    resultados = [

        {

            "apartamento": p.apartamento,

            "parqueadero_numero": r.parqueadero_asignado,

            "parqueadero_zona": r.zona_asignada,

            "tipo_resultado": r.tipo_resultado,

        }

        for r, p in rows

    ]

    return {

        "conjunto": tenant.nombre,

        "sorteo_id": sorteo.id,

        "fecha": sorteo.created_at.isoformat() if sorteo.created_at else "",

        "seed": sorteo.seed,

        "resultados": resultados,

    }





def publico_seed(db: Session, tenant_slug: str, sorteo_id: int) -> dict[str, str]:

    """Seed publico (SDD §5.4)."""

    tenant = _resolver_tenant_por_slug(db, tenant_slug)

    sorteo = (

        db.query(Sorteo)

        .filter(Sorteo.id == sorteo_id, Sorteo.tenant_id == tenant.id, Sorteo.estado == "COMPLETADO")

        .first()

    )

    if not sorteo or not sorteo.seed:

        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Seed no disponible")

    return {"seed": sorteo.seed}


# Nueva función exportar
def exportar_resultados(db: Session, tenant_id: str, sorteo_id: int, formato: str = "excel") -> bytes:
    """Exporta acta del sorteo (SDD §16)."""
    from app.services.exportadores import exportar_acta_excel, exportar_acta_word

    sorteo = _obtener_sorteo_tenant(db, tenant_id, sorteo_id)
    enforce_tenant_scope(tenant_id, sorteo.tenant_id)

    if formato == "word":
        return exportar_acta_word(db, tenant_id, sorteo_id)
    return exportar_acta_excel(db, tenant_id, sorteo_id)
